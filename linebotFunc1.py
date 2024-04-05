from flask import Flask, request, abort, jsonify

from linebot.exceptions import (
    InvalidSignatureError
)
from linebot import (
    LineBotApi, WebhookHandler,
)
from linebot.models import *

from linebot.models import (
    CarouselTemplate, CarouselColumn, TemplateSendMessage, PostbackAction, MessageAction
)
from linebot.exceptions import LineBotApiError
from datetime import datetime, timedelta
from openpyxl import load_workbook
import logging
import firebase_admin
from firebase_admin import credentials, initialize_app, storage

# 初始化 Firebase Admin
cred = credentials.Certificate("C:\\Users\\hcjarch\\PycharmProjects\\linebot1.0\\linebot.json")
firebase_app = firebase_admin.initialize_app(cred, {'storageBucket': 'loveless-b16d1.appspot.com'})

app = Flask(__name__)
# LINE BOT info
line_bot_api = LineBotApi('C265nOl8F1HIi81pF1BVCb6DIelXeSBttFPlrWNQoj0iG588dmODNJcwCpLwBwbdFEEFc1iy/o/6SsVGu2qt9+8f6hZzejnIxfUK1huUV3xjO9z2JblA8hgIRnXcu6PcgpxDNBvcdlM3oDo9iqYtDwdB04t89/1O/w1cDnyilFU=')
handler = WebhookHandler('88d2048c9f22572c4faddcfc8f198663')

user_data = {}

questions = [
    "一、比起真正擁有一段親密關係，你更容易進入對完美關係的幻想。",
    "二、當有人隨意地向你表現出浪漫的興趣，你會本能地想拒絕，推開這個人",
    "三、你感到你只是喜歡一段親密關係最開始的時候，一但有衝突的信號，你就想馬上切斷關係逃跑",
    "四、你會對於做出承諾感到焦慮，因為這會限制你的更多選擇與可能性。",
    "五、你會本能的，抵抗或回覆這個人的訊息、電話",
    "六、對於並不需要一直見面的人，你反而更容易在情感上親近對方",
    "七、在朋友、家庭、熟人面前,你有嚴格的區分,區分你自己應該呈現怎麼樣的樣子。在每種狀態下你都像是一個不同的人。",
    "八、當由於「太忙」,而無法對他人做出與交往相關的承諾(比如見面、約會)時,你感到驕傲或開心。",
    "九、你害怕當人們知道真實的你是什麼樣以後,就不會再喜歡你。",
    # 其餘的題目...
]

# Message event
@handler.add(MessageEvent, message=TextMessage)
def handle_text_message(event):
    user_id = event.source.user_id
    message = event.message.text
    reply_token = event.reply_token

    # 如果收到 '開始愛無能測驗' 命令，則初始化或重置用戶狀態並開始測驗
    if message == "開始愛無能測驗":
        user_data[user_id] = {'score': 0, 'question_index': 0}
        send_intro_message_and_first_question(reply_token, user_id)
        pass
    elif message == "預約參觀":
        handle_reservation_request(reply_token)
    # 如果用戶已經開始測驗，則處理他們的答案
    elif user_id in user_data:
        if message.isdigit() and 1 <= int(message) <= 3:
            handle_answer(reply_token, user_id, int(message))
        else:
            # 如果答案不是數字1, 2, 或 3，則提醒用戶輸入有效的答案
            line_bot_api.reply_message(
                reply_token,
                TextSendMessage(text="請輸入有效的命令或回答1, 2, 3中的一個。")
            )
    elif "預約時間" in message:
        handle_booking_confirmation(user_id, message, reply_token)
    # 如果用戶未在測驗中且發送其他命令，可以選擇忽略或提供不同的回應
    elif '.png' in message or '.jpg' in message:
        # Call the function to get the image URL from Firebase
        image_url = get_image_url_from_firebase(message)
        if image_url:
            # Send the image to the user
            line_bot_api.reply_message(
                reply_token,
                ImageSendMessage(original_content_url=image_url, preview_image_url=image_url)
            )
        else:
            line_bot_api.reply_message(
                reply_token,
                TextSendMessage(text='Image not found.')
            )
    else:
        # 這裡可以加入你想要的任何其他關鍵字回應
        # line_bot_api.reply_message(
        #     reply_token,
        #     TextSendMessage(text="你輸入的我不能理解，你可以試著說:預約參觀")
        # )
        pass

def send_intro_message_and_first_question(reply_token, user_id):
    intro_message = " 愛無能測驗介紹\n測驗說明與分數對應：\n1分：不符合。2分：有時符合。3分：完全符合。\n\n開始後請依序回答問題。範例回覆：只需輸入羅馬數字1、2或3\n\n*測驗一但開始，就需要完成完整的測驗問答!*"
    # 发送介绍信息
    line_bot_api.reply_message(
        reply_token,
        [TextSendMessage(text=intro_message), TextSendMessage(text=questions[0])]
    )
    # 更新问题索引为下一题
    user_data[user_id]['question_index'] = 1

def handle_answer(reply_token, user_id, answer):
    user_data[user_id]['score'] += answer
    # 不再在这里递增question_index，只在问问题后递增
    ask_question(reply_token, user_id)

def ask_question(reply_token, user_id):
    index = user_data[user_id]['question_index']
    if index < len(questions):
        line_bot_api.push_message(
            user_id,
            TextSendMessage(text=questions[index])
        )
        user_data[user_id]['question_index'] += 1  # 在这里递增
    else:
        calculate_result(reply_token, user_id)  # 确保传递reply_token

def calculate_result(reply_token, user_id):
    score = user_data[user_id]['score']
    result = ''
    if score <= 12:
        result = '完全沒有愛無能傾向'
    elif 13 <= score <= 17:
        result = '愛無能傾向低'
    elif 18 <= score <= 21:
        result = '愛無能傾向高'
    else:
        result = '愛無能者'

    # 使用 reply_token 发送结果消息
    line_bot_api.reply_message(
        reply_token,
        TextSendMessage(text=f'測驗結果: {score}分，{result}\n\n愛無能小百科 : \n愛無能指的是無法有效地表達、理解或體驗愛情的情感狀態。這可能源自個人的心理、情感或社交因素，導致在愛情關係中感到困惑、不安或無助。這種情況可能表現為難以建立親密關係、缺乏情感連結或對愛情的恐懼。愛無能並非缺乏愛情的能力，而是對於愛情的適應能力受到阻礙，需要更深入的理解和支持。')
    )

    # 删除用户数据，测试结束
    del user_data[user_id]

def is_slot_booked(ws, date, time):
    for row in ws.iter_rows(min_row=2):  # 假设第一行是标题行
        # print(row[1].value, date)
        # print(row[2].value, time)
        if row[1].value == date and row[2].value == time:
            return True
    return False

def handle_reservation_request(reply_token):
    excel_file_path = 'C:\\Users\\hcjarch\\PycharmProjects\\linebot1.0\\bookings.xlsx'
    wb = load_workbook(excel_file_path)
    ws = wb.active

    # 用于存储生成的 bubbles 的列表
    carousel_contents = []

    dates = ["2024-04-15", "2024-04-17", "2024-04-18", "2024-04-19"]
    times = ["10:00", "10:30", "11:00", "11:30", "13:00", "13:30", "14:00", "14:30", "15:00"]

    for date in dates:
        buttons = []
        for time in times:
            if is_slot_booked(ws, date, time):
                # 已经被预约的时间段
                print("Hello")
                # button_content = {
                #     "type": "text",
                #     "color": "#aaaaaa",
                #     "text": time,
                #     "size": 'md',
                #     "align": "center",
                #     "lineSpacing": "20px"
                # }
            else:
                print(date, time)
                # 还没有被预约的时间段
                button_content = {
                    "type": "button",
                    "action": {
                        "type": "message",
                        "label": time,
                        "text": f"預約時間 {date} {time}",
                        "color": "#cccccc",

                    }
                }
                buttons.append(button_content)
        bubble_content = {
            "type": "bubble",
            "header": {
                "type": "box",
                "layout": "vertical",
                "contents": [
                    {
                        "type": "text",
                        "text": date,  # 使用循环中的日期变量
                        "weight": "bold",
                        "size": "xl"
                    },
                    {
                        "type": "box",
                        "layout": "vertical",
                        "margin": "lg",
                        "spacing": "sm",
                        "contents": [
                            {
                                "type": "box",
                                "layout": "baseline",
                                "spacing": "sm",
                                "contents": [
                                    {
                                        "type": "text",
                                        "text": "一個場次僅提供1人遊玩",
                                        "wrap": True,
                                        "color": "#666666",
                                        "size": "sm",
                                        "flex": 5
                                    }
                                ]
                            },
                            {
                                "type": "box",
                                "layout": "baseline",
                                "spacing": "sm",
                                "contents": [
                                    {
                                        "type": "text",
                                        "text": "體驗時間15~20分鐘",
                                        "wrap": True,
                                        "color": "#666666",
                                        "size": "sm",
                                        "flex": 5
                                    }
                                ]
                            }
                        ]
                    }
                ]
            },
            "body": {
                "type": "box",
                "layout": "vertical",
                "spacing": "sm",
                "contents": buttons,
            }
        }
        carousel_contents.append(bubble_content)

    flex_message = FlexSendMessage(
        alt_text='預約時段',
        contents={
            "type": "carousel",
            "contents": carousel_contents
        }
    )

    line_bot_api.reply_message(reply_token, flex_message)

def get_user_profile(user_id):
    try:
        profile = line_bot_api.get_profile(user_id)
        return profile.display_name
    except LineBotApiError as e:
        print(f"Error occurred: {e}")
        return None
def handle_booking_confirmation(user_id, message_text,reply_token):
    logging.basicConfig(level=logging.INFO)
    parts = message_text.split()  # ['預約時間', '4/15', '11:00']
    if len(parts) == 3 and parts[0] == '預約時間':
        try:
            booking_date = parts[1].split('(')[0]  # '4/15' without the day
            booking_time = parts[2]  # '11:00'
            booking_datetime = datetime.strptime(f'2024/{booking_date} {booking_time}', '%Y/%m/%d %H:%M')

            # 加载Excel工作簿
            # 用您的 Excel 文件的实际路径替换此处
            excel_file_path = 'C:\\Users\\hcjarch\\PycharmProjects\\linebot1.0\\bookings.xlsx'

            wb = load_workbook(excel_file_path)
            ws = wb.active
            if is_slot_booked(ws, booking_datetime.strftime('%Y-%m-%d'), booking_time):
                # If booked, send a message to the user
                line_bot_api.reply_message(
                    reply_token,
                    TextSendMessage(text=f"對不起，時間 {parts[1]} {booking_time} 已經被預約了。")
                )
                return
            display_name = get_user_profile(user_id)
            if display_name:
                name_to_write = display_name
            else:
                name_to_write = user_id  # 如果无法获取用户昵称，退回到 user_id
            # 找到表格的最后一行以便添加新预约
            last_row = ws.max_row + 1

            # 在表格中添加用户 ID、预约日期和时间
            ws.cell(row=last_row, column=1, value=name_to_write)  # 使用 name_to_write 替代 user_id
            ws.cell(row=last_row, column=2, value=booking_datetime.strftime('%Y-%m-%d'))
            ws.cell(row=last_row, column=3, value=booking_datetime.strftime('%H:%M'))

            # 保存工作簿
            wb.save(excel_file_path)

            confirmation_message = f"預約成功!您預約的時間為{booking_date} {booking_time}"
            line_bot_api.reply_message(reply_token, TextSendMessage(text=confirmation_message))
            logging.info(f'用户 {user_id} 预约了 {booking_date} {booking_time} 的时间')

        except ValueError as e:
            # Handle the ValueError if the date format is incorrect
            print("Date format error:", e)
            # You would handle the error here, possibly sending a message back to the user
            return

        except PermissionError as e:
            logging.error("无法保存文件，可能是因为权限不足或文件正在被其他程序使用。", exc_info=True)
            line_bot_api.push_message(
                user_id,
                TextSendMessage(text='无法保存预约信息，请确保文件不被其他程序使用，并且您有足够的权限。')
            )

        except Exception as e:
            logging.error("更新预约时发生错误: ", exc_info=True)
            line_bot_api.push_message(
                user_id,
                TextSendMessage(text='处理您的预约时发生错误，请重试。')
            )
        except ValueError as e:
            # Handle the ValueError if the date format is incorrect
            print("Date format error:", e)
            line_bot_api.reply_message(
                reply_token,
                TextSendMessage(text="日期格式錯誤，請使用正確的格式，例如：'預約時間 4/15 11:00'")
            )
            return
    else:
        # 回复用户错误的预约格式
        line_bot_api.push_message(
            user_id,
            TextSendMessage(text='无效的预约格式。请使用正确的格式，例如："預約時間 4/15 11:00"')
        )
    line_name = get_user_profile(user_id)
    if line_name is not None:
        # 如果成功获取昵称，使用昵称作为姓名存储
        ws.cell(row=last_row, column=1, value=line_name)
    else:
        # 如果获取昵称失败，使用 user_id
        ws.cell(row=last_row, column=1, value=user_id)

def get_image_url_from_firebase(file_name):
    # Make sure firebase_admin has been initialized with the proper credentials
    if not firebase_admin._apps:
        cred = credentials.Certificate('path/to/your/firebase/credentials.json')
        firebase_admin.initialize_app(cred, {'storageBucket': 'your-bucket-name.appspot.com'})

    bucket = storage.bucket()
    blob = bucket.blob('images/' + file_name)

    # Check if the file exists
    if blob.exists():
        # Generate a signed URL for the blob that expires after a specified time
        return blob.generate_signed_url(timedelta(minutes=15), method='GET')
    else:
        return None

@app.route('/get_image', methods=['POST'])
def get_image():
    data = request.json
    file_name = data.get('file_name')

    if file_name:
        # 构建对 Firebase Storage 中文件的引用
        bucket = storage.bucket()
        blob = bucket.blob(file_name)

        # 创建一个下载链接
        download_url = blob.generate_signed_url(timedelta(seconds=300), method='GET')

        # 返回下载链接
        return jsonify({'url': download_url})

    return 'File name is missing', 400

@app.route("/callback", methods=['POST'])
def callback():
    # get X-Line-Signature header value
    signature = request.headers['X-Line-Signature']

    # get request body as text
    body = request.get_data(as_text=True)
    app.logger.info("Request body: " + body)
    print(body)

    # handle webhook body
    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        abort(400)
    return 'OK'
if __name__ == "__main__":
    app.run()